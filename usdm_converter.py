import os
import re
import sys
import json
import functools
import urllib.request
import urllib.parse
import urllib.error

import openpyxl
from usdm4.api.code import Code
from usdm4.api.alias_code import AliasCode
from usdm4.api.activity import Activity
from usdm4.api.encounter import Encounter
from usdm4.api.biomedical_concept import BiomedicalConcept
from usdm4.api.scheduled_instance import ScheduledActivityInstance
from usdm4.api.schedule_timeline import ScheduleTimeline
from usdm4.api.eligibility_criterion import EligibilityCriterion, EligibilityCriterionItem
from usdm4.api.endpoint import Endpoint
from usdm4.api.objective import Objective
from usdm4.api.study_arm import StudyArm
from usdm4.api.study_intervention import StudyIntervention
from usdm4.api.population_definition import StudyDesignPopulation
from usdm4.api.study_epoch import StudyEpoch
from usdm4.api.study_element import StudyElement
from usdm4.api.study_cell import StudyCell
from usdm4.api.timing import Timing
from usdm4.api.study import Study
from usdm4.api.study_version import StudyVersion
from usdm4.api.study_design import InterventionalStudyDesign
from usdm4.api.wrapper import Wrapper
from usdm4.api.identifier import StudyIdentifier
from usdm4.api.study_title import StudyTitle
from usdm4.api.organization import Organization


EVS_SEARCH_URL = "https://api-evsrest.nci.nih.gov/api/v1/concept/ncit/search"

_OBJECTIVE_LEVEL_CT = {
    "primary": ("C85826", "Trial Primary Objective"),
    "secondary": ("C85827", "Trial Secondary Objective"),
    "exploratory": ("C163559", "Trial Exploratory Objective"),
}
_ENDPOINT_LEVEL_CT = {
    "primary": ("C94496", "Primary Endpoint"),
    "secondary": ("C139173", "Secondary Endpoint"),
    "exploratory": ("C170559", "Exploratory Endpoint"),
}
_DATA_ORIGIN_TYPE_CT = ("C188866", "Data Generated Within Study")
_ENCOUNTER_TYPE_CT = ("C25716", "Visit")
_ORGANIZATION_TYPE_CT = ("C54149", "Drug Company")


def _build_ct_code(ids, code, decode, codelist_label):
    return Code(
        id=ids.next("CODE"),
        code=code,
        codeSystem=f"CDISC DDF {codelist_label} codelist",
        codeSystemVersion="2025-09-26",
        decode=decode,
    )


@functools.lru_cache(maxsize=None)
def _lookup_ncit_code(term):
    params = urllib.parse.urlencode({"term": term, "type": "match", "pageSize": 5})
    url = f"{EVS_SEARCH_URL}?{params}"
    try:
        with urllib.request.urlopen(url, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        return None, None
    concepts = data.get("concepts", [])
    for c in concepts:
        if c.get("name", "").strip().lower() == term.strip().lower():
            return c["code"], c.get("version")
    if concepts:
        return concepts[0]["code"], concepts[0].get("version")
    return None, None


_ENCOUNTER_TYPE_KEYWORDS = [
    "Premature Discontinuation", "Early Termination", "End of Treatment",
    "End of Study", "Safety Follow-Up", "Follow-Up", "Follow up",
    "Re-Induction", "Induction", "Randomization", "Maintenance",
    "Screening", "Baseline", "Run-in", "Washout", "Unscheduled",
    "Extension", "Discontinuation",
]


def _lookup_encounter_type_code(label):
    code, version = _lookup_ncit_code(label)
    if code:
        return code, version
    lower = label.lower()
    for keyword in _ENCOUNTER_TYPE_KEYWORDS:
        if keyword.lower() in lower:
            code, version = _lookup_ncit_code(keyword)
            if code:
                return code, version
    return None, None


def _lookup_activity_code(label):
    code, version = _lookup_ncit_code(label)
    if code:
        return code, version
    paren_idx = label.find("(")
    if paren_idx > 0:
        trimmed = label[:paren_idx].strip().rstrip(",")
        if trimmed and trimmed != label:
            code, version = _lookup_ncit_code(trimmed)
            if code:
                return code, version
    return None, None


def _build_code(ids, term, unmapped_terms):
    code, version = _lookup_ncit_code(term)
    if code:
        return Code(
            id=ids.next("CODE"),
            code=code,
            codeSystem="NCI Thesaurus (via NCI EVS, CDISC's terminology partner)",
            codeSystemVersion=version or "unknown",
            decode=term,
        )
    unmapped_terms.add(term)
    return Code(
        id=ids.next("CODE"),
        code=_slug(term).upper(),
        codeSystem="Sponsor Defined (no NCIt match found)",
        codeSystemVersion="1.0",
        decode=term,
    )


_WEEK_RE = re.compile(r"\bWeek\)?\s*(-?\d+)\b", re.I)
_DAY_RE = re.compile(r"\bDay\)?\s*(-?\d+)\b", re.I)
_N_DAY_RE = re.compile(r"\b(-?\d+)-Day\b", re.I)


def _parse_visit_offset(label):
    m = _DAY_RE.search(label)
    if m:
        return int(m.group(1)), m.group(0)
    m = _N_DAY_RE.search(label)
    if m:
        return int(m.group(1)), m.group(0)
    m = _WEEK_RE.search(label)
    if m:
        return int(m.group(1)) * 7, m.group(0)
    return None, None


def convert_eligibility(elig_data, ids, unmapped_terms):
    criteria, items = [], []

    def _add(category_term, prefix, texts):
        for i, text in enumerate(texts, start=1):
            item = EligibilityCriterionItem(
                id=ids.next("CRITITEM"),
                name=f"{prefix}{i}",
                text=text,
            )
            items.append(item)
            criteria.append(EligibilityCriterion(
                id=ids.next("CRIT"),
                name=f"{prefix}{i}",
                category=_build_code(ids, category_term, unmapped_terms),
                identifier=f"{prefix}{i}",
                criterionItemId=item.id,
            ))

    _add("Inclusion Criteria", "INC", elig_data.get("inclusionCriteria", []))
    _add("Exclusion Criteria", "EXC", elig_data.get("exclusionCriteria", []))

    population = StudyDesignPopulation(
        id=ids.next("POP"),
        name="Study Population",
        includesHealthySubjects=False,
        criterionIds=[c.id for c in criteria],
    )
    return criteria, items, population


def _objective_level_code(ids, level_term, unmapped_terms):
    match = _OBJECTIVE_LEVEL_CT.get(level_term.strip().lower())
    if match:
        return _build_ct_code(ids, match[0], match[1], "Objective Level")
    unmapped_terms.add(f"Objective level '{level_term}' (no match in Primary/Secondary/Exploratory)")
    return _build_code(ids, level_term, unmapped_terms)


def _endpoint_level_code(ids, level_term, unmapped_terms):
    match = _ENDPOINT_LEVEL_CT.get(level_term.strip().lower())
    if match:
        return _build_ct_code(ids, match[0], match[1], "Endpoint Level")
    unmapped_terms.add(f"Endpoint level '{level_term}' (no match in Primary/Secondary/Exploratory)")
    return _build_code(ids, level_term, unmapped_terms)


def convert_objectives(objectives_data, ids, unmapped_terms):
    objectives = []
    level_counters = {}
    for o in objectives_data:
        level_term = o.get("level") or "Unspecified"
        level_counters[level_term] = level_counters.get(level_term, 0) + 1
        obj_name = f"{level_term} Objective {level_counters[level_term]}"

        endpoints = []
        for j, e in enumerate(o.get("endpoints", []), start=1):
            desc = e.get("description", "")
            endpoints.append(Endpoint(
                id=ids.next("EP"),
                name=f"{obj_name} - Endpoint {j}",
                text=desc,
                purpose=desc,
                level=_endpoint_level_code(ids, level_term, unmapped_terms),
            ))

        objectives.append(Objective(
            id=ids.next("OBJ"),
            name=obj_name,
            text=o.get("description", ""),
            level=_objective_level_code(ids, level_term, unmapped_terms),
            endpoints=endpoints,
        ))
    return objectives


def convert_arms(arms_data, ids, unmapped_terms):
    study_arms = []
    interventions = []
    arm_intervention_ids = []
    for a in arms_data:
        type_term = a.get("type") or "Experimental"
        type_code = _build_code(ids, type_term, unmapped_terms)

        iv_list = a.get("interventions", [])
        iv_summary = "; ".join(
            iv["name"] + (
                " (" + ", ".join(filter(None, [iv.get("dose"), iv.get("route"), iv.get("frequency")])) + ")"
                if any([iv.get("dose"), iv.get("route"), iv.get("frequency")]) else ""
            )
            for iv in iv_list if iv.get("name")
        )
        description = a.get("description", "") or a["name"]
        if iv_summary:
            description = f"{description}\n\nInterventions: {iv_summary}"

        study_arms.append(StudyArm(
            id=ids.next("ARM"),
            name=a["name"],
            label=a["name"],
            description=description,
            type=type_code,
            dataOriginDescription="Data collected from enrolled study subjects",
            dataOriginType=_build_ct_code(ids, _DATA_ORIGIN_TYPE_CT[0], _DATA_ORIGIN_TYPE_CT[1], "Study Arm Data Origin Type"),
        ))

        role_term = "Comparator" if re.search(r"placebo|comparator", type_term, re.I) else "Investigational Product"
        this_arm_iv_ids = []
        for iv in iv_list:
            if not iv.get("name"):
                continue
            iv_desc = ", ".join(filter(None, [iv.get("dose"), iv.get("route"), iv.get("frequency")])) or iv["name"]
            new_iv = StudyIntervention(
                id=ids.next("IV"),
                name=iv["name"],
                label=iv["name"],
                description=iv_desc,
                role=_build_code(ids, role_term, unmapped_terms),
                type=_build_code(ids, iv.get("type") or "Drug", unmapped_terms),
            )
            interventions.append(new_iv)
            this_arm_iv_ids.append(new_iv.id)
        arm_intervention_ids.append(this_arm_iv_ids)
    return study_arms, interventions, arm_intervention_ids


def convert_design_matrix(study_arms, arm_intervention_ids, timelines, ids, unmapped_terms):
    epochs = []
    for tl in timelines:
        epoch_name = tl.name or tl.label
        ecode, eversion = _lookup_encounter_type_code(epoch_name)
        if ecode:
            type_code = Code(
                id=ids.next("CODE"), code=ecode,
                codeSystem="NCI Thesaurus (via NCI EVS, CDISC's terminology partner)",
                codeSystemVersion=eversion or "unknown", decode=epoch_name,
            )
        else:
            unmapped_terms.add(epoch_name)
            type_code = Code(
                id=ids.next("CODE"), code=_slug(epoch_name).upper(),
                codeSystem="Sponsor Defined (no NCIt match found)",
                codeSystemVersion="1.0", decode=epoch_name,
            )
        epochs.append(StudyEpoch(id=ids.next("EPOCH"), name=epoch_name, label=epoch_name, type=type_code))

    elements, study_cells, unmatched_arms = [], [], []
    for arm, iv_ids in zip(study_arms, arm_intervention_ids):
        element = StudyElement(
            id=ids.next("ELEM"), name=f"{arm.name} Element", label=f"{arm.name} Element",
            studyInterventionIds=iv_ids,
        )
        elements.append(element)

        matched_epoch = None
        for kw in _ENCOUNTER_TYPE_KEYWORDS:
            if kw.lower() in arm.name.lower():
                matched_epoch = next((e for e in epochs if kw.lower() in (e.name or "").lower()), None)
                if matched_epoch:
                    break

        if matched_epoch:
            study_cells.append(StudyCell(id=ids.next("CELL"), armId=arm.id, epochId=matched_epoch.id, elementIds=[element.id]))
        else:
            unmatched_arms.append(arm.name)

    return epochs, elements, study_cells, unmatched_arms


def build_wrapper(study_name, ids, timelines, all_activities, all_encounters, all_biomedical_concepts,
                   eligibility_criteria, eligibility_items, population, study_arms, study_interventions,
                   objectives, epochs, elements, study_cells, unmapped_terms):
    if population is None:
        population = StudyDesignPopulation(
            id=ids.next("POP"),
            name="Study Population",
            includesHealthySubjects=False,
            criterionIds=[],
        )

    study_design = InterventionalStudyDesign(
        id=ids.next("SD"),
        name="Study Design",
        rationale="Not extracted from protocol text -- placeholder pending manual entry.",
        model=_build_code(ids, "Parallel Study", unmapped_terms),
        arms=study_arms,
        studyCells=study_cells,
        epochs=epochs,
        elements=elements,
        encounters=all_encounters,
        activities=all_activities,
        objectives=objectives,
        population=population,
        scheduleTimelines=timelines,
        eligibilityCriteria=eligibility_criteria,
        studyInterventionIds=[iv.id for iv in study_interventions],
    )

    sponsor_org = Organization(
        id=ids.next("ORG"),
        name="Sponsor (not extracted from protocol text)",
        type=_build_ct_code(ids, _ORGANIZATION_TYPE_CT[0], _ORGANIZATION_TYPE_CT[1], "Organization Type"),
        identifierScheme="Not extracted from protocol text -- placeholder pending manual entry.",
        identifier="UNKNOWN",
        managedSites=[],
    )
    study_identifier = StudyIdentifier(
        id=ids.next("SI"),
        text="Not extracted from protocol text -- placeholder pending manual entry.",
        scopeId=sponsor_org.id,
    )
    study_title = StudyTitle(
        id=ids.next("STITLE"),
        text=study_name,
        type=_build_code(ids, "Official Study Title", unmapped_terms),
    )

    study_version = StudyVersion(
        id=ids.next("SV"),
        versionIdentifier="1.0",
        rationale="Not extracted from protocol text -- placeholder pending manual entry.",
        studyDesigns=[study_design],
        eligibilityCriterionItems=eligibility_items,
        studyInterventions=study_interventions,
        biomedicalConcepts=all_biomedical_concepts,
        organizations=[sponsor_org],
        studyIdentifiers=[study_identifier],
        titles=[study_title],
    )

    study = Study(
        name=study_name,
        versions=[study_version],
    )

    return Wrapper(
        study=study,
        usdmVersion="4.0.0",
        systemName="usdm_converter.py",
        systemVersion="1.0",
    )


class _IdGen:
    def __init__(self):
        self._counters = {}

    def next(self, prefix):
        self._counters[prefix] = self._counters.get(prefix, 0) + 1
        return f"{prefix}_{self._counters[prefix]}"


def _slug(text):
    return re.sub(r"[^a-z0-9]+", "_", text.strip().lower()).strip("_") or "unnamed"


def _read_sheet_table(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 2:
        return None, [], []
    title = rows[0][0]
    raw_header = rows[1][1:]
    header_positions = [i for i, c in enumerate(raw_header) if c is not None]
    header = [str(raw_header[i]).strip() for i in header_positions]
    data_rows = []
    for r in rows[2:]:
        activity = r[0]
        if activity is None or str(activity).strip() in ("", "Footnotes"):
            break
        raw_marks = r[1:]
        marks = [raw_marks[i] if i < len(raw_marks) else None for i in header_positions]
        data_rows.append((str(activity).strip(), marks))
    return title, header, data_rows


def convert_sheet(ws, ids, unmapped_labels, unmapped_activity_labels, unmapped_terms, unmapped_timing_labels):
    title, header, data_rows = _read_sheet_table(ws)
    if not header or not data_rows:
        return None

    encounters = []
    encounter_ids = []
    anchor_instances = []
    timings = []
    for col_label in header:
        code = _build_ct_code(ids, _ENCOUNTER_TYPE_CT[0], _ENCOUNTER_TYPE_CT[1], "Encounter Type")
        enc_id = ids.next("ENC")
        encounters.append(
            Encounter(id=enc_id, name=col_label, label=col_label, type=code)
        )
        encounter_ids.append(enc_id)

        anchor_id = ids.next("SAI")
        anchor_instances.append(
            ScheduledActivityInstance(
                id=anchor_id, name=f"Visit anchor: {col_label}",
                activityIds=[], encounterId=enc_id,
            )
        )

        days, matched_text = _parse_visit_offset(col_label)
        if days is not None:
            timings.append(Timing(
                id=ids.next("TIMING"),
                name=f"Timing for {col_label}",
                type=_build_code(ids, "Fixed Reference", unmapped_terms),
                value=str(days),
                valueLabel=matched_text,
                relativeToFrom=_build_code(ids, "Start to Start", unmapped_terms),
                relativeFromScheduledInstanceId=anchor_id,
            ))
        else:
            unmapped_timing_labels.add(col_label)

    activities = []
    biomedical_concepts = []
    activity_instances = []
    for activity_label, marks in data_rows:
        act_id = ids.next("ACT")
        act_ncit_code, act_ncit_version = _lookup_activity_code(activity_label)
        bc_ids = []
        if act_ncit_code:
            standard_code = Code(
                id=ids.next("CODE"),
                code=act_ncit_code,
                codeSystem="NCI Thesaurus (via NCI EVS, CDISC's terminology partner)",
                codeSystemVersion=act_ncit_version or "unknown",
                decode=activity_label,
            )
            alias_code = AliasCode(
                id=ids.next("ALIASCODE"),
                standardCode=standard_code,
                standardCodeAliases=[],
            )
            bc = BiomedicalConcept(
                id=ids.next("BC"),
                name=activity_label,
                label=activity_label,
                synonyms=[],
                reference=f"NCIt:{act_ncit_code}",
                properties=[],
                code=alias_code,
                notes=[],
            )
            biomedical_concepts.append(bc)
            bc_ids = [bc.id]
        else:
            unmapped_activity_labels.add(activity_label)
        activities.append(
            Activity(
                id=act_id,
                name=activity_label,
                label=activity_label,
                biomedicalConceptIds=bc_ids,
            )
        )
        for col_idx, mark in enumerate(marks):
            if mark is None or str(mark).strip() == "":
                continue
            sai_id = ids.next("SAI")
            activity_instances.append(
                ScheduledActivityInstance(
                    id=sai_id,
                    name=sai_id,
                    activityIds=[act_id],
                    encounterId=encounter_ids[col_idx],
                )
            )

    if not activity_instances:
        return None

    all_instances = activity_instances + anchor_instances
    timeline = ScheduleTimeline(
        id=ids.next("TL"),
        name=title or ws.title,
        label=title or ws.title,
        mainTimeline=False,
        entryCondition=f"Start of {title or ws.title}",
        entryId=activity_instances[0].id,
        instances=all_instances,
        timings=timings,
    )
    return timeline, activities, encounters, biomedical_concepts, timings


def convert_workbook(xlsx_path, protocol_json_path=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ids = _IdGen()
    timelines, all_activities, all_encounters, all_biomedical_concepts, all_timings = [], [], [], [], []
    skipped_sheets = []
    unmapped_labels = set()
    unmapped_activity_labels = set()
    unmapped_timing_labels = set()
    unmapped_protocol_terms = set()

    for ws in wb.worksheets:
        result = convert_sheet(ws, ids, unmapped_labels, unmapped_activity_labels, unmapped_protocol_terms, unmapped_timing_labels)
        if result is None:
            skipped_sheets.append(ws.title)
            continue
        timeline, activities, encounters, biomedical_concepts, timings = result
        timelines.append(timeline)
        all_activities.extend(activities)
        all_encounters.extend(encounters)
        all_biomedical_concepts.extend(biomedical_concepts)
        all_timings.extend(timings)

    if timelines:
        timelines[0].mainTimeline = True

    eligibility_criteria, eligibility_items, population = [], [], None
    objectives, study_arms, study_interventions, arm_intervention_ids = [], [], [], []
    epochs, elements, study_cells, unmatched_arms = [], [], [], []
    if protocol_json_path and os.path.exists(protocol_json_path):
        with open(protocol_json_path) as f:
            protocol_data = json.load(f)
        eligibility_criteria, eligibility_items, population = convert_eligibility(
            protocol_data.get("eligibilityCriteria", {}), ids, unmapped_protocol_terms
        )
        objectives = convert_objectives(protocol_data.get("objectives", []), ids, unmapped_protocol_terms)
        study_arms, study_interventions, arm_intervention_ids = convert_arms(
            protocol_data.get("studyArms", []), ids, unmapped_protocol_terms
        )
        if study_arms:
            epochs, elements, study_cells, unmatched_arms = convert_design_matrix(
                study_arms, arm_intervention_ids, timelines, ids, unmapped_protocol_terms
            )

    doc = {
        "_conversionNotes": (
            "Generated by usdm_converter.py from soa_detector.py's Excel output. "
            f"Sheets skipped (no usable table found): {skipped_sheets or 'none'}. "
            f"Activity labels with no NCIt match (no biomedicalConceptIds): "
            f"{sorted(unmapped_activity_labels) or 'none'}. "
            f"Encounter labels with no parseable Day/Week offset (no Timing "
            f"built): {sorted(unmapped_timing_labels) or 'none'}. "
            + (
                "protocol_data.json was not found, so eligibilityCriteria, "
                "studyArms, objectives, epochs, elements, and studyCells below "
                "are empty -- this run only covers the SoA."
                if not protocol_json_path or not os.path.exists(protocol_json_path) else
                "Arms with no epoch keyword match (no StudyCell built): "
                f"{sorted(unmatched_arms) or 'none'}. "
                f"Eligibility/objective/arm/epoch terms with no NCIt match: "
                f"{sorted(unmapped_protocol_terms) or 'none'}."
            )
        ),
        "scheduleTimelines": [json.loads(t.to_json()) for t in timelines],
        "activities": [json.loads(a.to_json()) for a in all_activities],
        "encounters": [json.loads(e.to_json()) for e in all_encounters],
        "biomedicalConcepts": [json.loads(b.to_json()) for b in all_biomedical_concepts],
        "timings": [json.loads(t.to_json()) for t in all_timings],
        "eligibilityCriteria": [json.loads(c.to_json()) for c in eligibility_criteria],
        "eligibilityCriterionItems": [json.loads(i.to_json()) for i in eligibility_items],
        "population": json.loads(population.to_json()) if population else None,
        "studyArms": [json.loads(a.to_json()) for a in study_arms],
        "studyInterventions": [json.loads(i.to_json()) for i in study_interventions],
        "objectives": [json.loads(o.to_json()) for o in objectives],
        "epochs": [json.loads(e.to_json()) for e in epochs],
        "elements": [json.loads(e.to_json()) for e in elements],
        "studyCells": [json.loads(c.to_json()) for c in study_cells],
    }

    study_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    wrapper = build_wrapper(
        study_name, ids, timelines, all_activities, all_encounters, all_biomedical_concepts,
        eligibility_criteria, eligibility_items, population, study_arms, study_interventions,
        objectives, epochs, elements, study_cells, unmapped_protocol_terms,
    )

    return doc, wrapper


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else input("Path to soa_tables.xlsx: ").strip()
    default_protocol_json = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)), "protocol_data.json")
    protocol_json_path = sys.argv[2] if len(sys.argv) > 2 else default_protocol_json

    doc, wrapper = convert_workbook(xlsx_path, protocol_json_path)
    out_path = re.sub(r"\.xlsx$", "", xlsx_path) + "_usdm.json"
    with open(out_path, "w") as f:
        json.dump(doc, f, indent=2)

    wrapper_path = re.sub(r"\.xlsx$", "", xlsx_path) + "_usdm_wrapper.json"
    with open(wrapper_path, "w") as f:
        f.write(wrapper.to_json())
    print(f"Timelines: {len(doc['scheduleTimelines'])}")
    print(f"Activities: {len(doc['activities'])}")
    print(f"Encounters: {len(doc['encounters'])}")
    total_sais = sum(len(t["instances"]) for t in doc["scheduleTimelines"])
    print(f"ScheduledActivityInstances (marked cells): {total_sais}")
    mapped = sum(
        1 for e in doc["encounters"]
        if "Encounter Type" in e["type"]["codeSystem"]
    )
    print(f"Encounters with a valid DDF Encounter Type code (generic 'Visit' -- "
          f"see docstring in convert_sheet): {mapped} / {len(doc['encounters'])}")
    activities_with_bc = sum(1 for a in doc["activities"] if a["biomedicalConceptIds"])
    print(f"Activities with real NCIt-based BiomedicalConcepts: {activities_with_bc} / {len(doc['activities'])}")
    print(f"BiomedicalConcepts created: {len(doc['biomedicalConcepts'])}")
    print(f"Timings built: {len(doc['timings'])} / {len(doc['encounters'])} encounters "
          f"(rest had no parseable Day/Week offset)")
    if os.path.exists(protocol_json_path):
        print(f"Eligibility criteria: {len(doc['eligibilityCriteria'])}")
        print(f"Study arms: {len(doc['studyArms'])}")
        print(f"Study interventions: {len(doc['studyInterventions'])}")
        print(f"Objectives: {len(doc['objectives'])}")
        total_endpoints = sum(len(o["endpoints"]) for o in doc["objectives"])
        print(f"Endpoints: {total_endpoints}")
        print(f"Epochs: {len(doc['epochs'])}")
        print(f"Elements: {len(doc['elements'])}")
        print(f"StudyCells (arms formally linked to an epoch): {len(doc['studyCells'])} / {len(doc['studyArms'])}")
    else:
        print(f"(No protocol_data.json found at {protocol_json_path} -- "
              f"eligibility/arms/objectives not included. Run "
              f"protocol_extractor.py first if you want those.)")
    print(f"Saved debug view to {out_path}")
    print(f"Saved real, validatable USDM document to {wrapper_path} -- validate THIS file, not the debug one above.")