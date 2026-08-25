import os
import re
import json
import difflib
import fitz
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from google import genai
from google.genai import types

client = genai.Client(api_key=os.environ.get("GEMINI_API_KEY"))

STRUCTURE_SCHEMA = {
    "type": "object",
    "properties": {
        "columns": {"type": "array", "items": {"type": "string"}},
        "activities": {"type": "array", "items": {"type": "string"}},
        "footnotes": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {"marker": {"type": "string"}, "text": {"type": "string"}},
                "required": ["marker", "text"],
            },
        },
    },
    "required": ["columns", "activities", "footnotes"],
}

STRUCTURE_INSTRUCTIONS = """Look at this Schedule of Activities table. Do NOT read any checkmarks or
in-grid cell content yet — for this pass, only identify:

- "columns": the visit/timepoint column headers, left to right, excluding the Activity/procedure
  column. Merge multi-row/nested headers into one string per real column (e.g. a "Week 1" group
  header spanning "Day 1", "Day 2" sub-columns becomes "Week 1 - Day 1", "Week 1 - Day 2"). The
  number of strings must equal the number of real vertical columns actually drawn.
- "activities": the exact activity/procedure row labels, top to bottom, exactly as written
  (including footnote markers like "^a"). Exclude section/category divider rows that aren't
  themselves a procedure (e.g. "INTERVIEWS & QUESTIONNAIRES" is a section header, not an activity).
- "footnotes": any footnote markers (a, b, c, *, etc.) referenced from activity labels or column
  headers, and their full definition text (usually printed below the table). Empty list if none.

Respond with JSON only, matching the schema."""


MARKS_SCHEMA = {
    "type": "object",
    "properties": {
        "rows": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "activity": {"type": "string"},
                    "marked_cells": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {"column": {"type": "string"}, "value": {"type": "string"}},
                            "required": ["column", "value"],
                        },
                    },
                },
                "required": ["activity", "marked_cells"],
            },
        },
    },
    "required": ["rows"],
}


def build_marks_instructions(columns_subset, all_columns, activities):
    other_columns = [c for c in all_columns if c not in columns_subset]
    return f"""This table's full structure has already been confirmed:
All activities, top to bottom: {json.dumps(activities)}
All columns, left to right: {json.dumps(all_columns)}

For THIS pass, you only need to report marks for this SUBSET of columns:
{json.dumps(columns_subset)}

The table also has these other columns, which you should completely ignore for this pass — do not
report anything for them, even if a row's marks visually continue into them: {json.dumps(other_columns)}

Go through the table ONE ROW AT A TIME, in the exact order of the activities list above. For each
activity, look ONLY at the columns listed in the subset above, left to right within that subset,
checking each one individually for a mark before moving to the next. Restricting your attention to
fewer columns per row means you can check each one carefully instead of scanning the whole row at
once. Do not skip ahead or infer a cell's contents from neighboring rows or columns.

Some tables have two or more rows with closely related labels right next to each other (e.g. a
procedure and then a separate row for "Automatic calculation of [that same procedure] in EDC", or a
parent item followed by one of its own components). These are still separate, independent rows —
each has its own row of grid cells and can be marked completely differently from its neighbor. Treat
every row as its own independent visual scan; never assume, copy, or reuse a neighboring row's marks
just because the two labels are topically related or visually close together.

If a cell's content spans multiple lines (e.g. "0 (pre-dose)" on one line and "2, 4, 6, 8 hours
post-dose" on the next), "value" must include every line, in reading order, joined with a single
space — never just the first line. Copy the complete cell text, not a truncated summary of it.

There are two different kinds of cells, and they need different handling:
- Simple checkmark/X/shaded-fill cells, where it can genuinely be ambiguous whether a faint mark is
  there at all — for these, if you're not confident, leave the cell out rather than guess.
- Cells containing legible printed text (times, doses, descriptions, etc.) — for these there is no
  ambiguity about whether something is there, only about transcribing it correctly. If a cell
  visibly has any text in it, it MUST appear in "marked_cells" with that text transcribed in full.
  Never drop a text-bearing cell just because part of it was hard to read — report your best
  complete reading instead of omitting it.

Before responding, re-scan each row of the subset one more time, left to right, and confirm every
column that visibly contains a checkmark or any text is represented in "marked_cells" for that row.

Respond with JSON: "rows", one entry per activity, in the exact order and exact spelling given above.
Each entry has "activity" (copied verbatim from the activities list) and "marked_cells" — the cells
you confirmed are marked, restricted to the subset columns above, each with "column" (copied verbatim
from the subset list) and "value" (the checkmark/text actually shown in that cell, complete and
unabridged). Only leave a cell out if it is genuinely blank, or if it's a checkmark-style cell you're
truly unsure about — never omit a cell that visibly contains text."""


def detect_soas(pdf_path):
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[{"parts": [
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_bytes}},
            {"text": """Find every Schedule of Activities / Assessments / Blood Collection / Biomarker
table in this clinical trial protocol.

A real match is a grid where each ROW is a procedure/activity/sample (e.g. "Vital Signs",
"Informed Consent", "PK Blood Draw") and each COLUMN is a study visit or timepoint (e.g. "Screening",
"Day 1", "Cycle 2 Day 15", "Follow-Up") — the cells show whether/when that row's activity happens at
that column's visit, usually via a checkmark, X, or shaded fill.

The same set of visit/timepoint columns must be shared across most or all of the rows — that shared
grid is what makes it a schedule. If a table's "columns" are really just per-item facts about that
one specific row (e.g. one medication's own Dose/Start/Stop), and a different row would have
entirely different, unrelated column values, that is NOT a visit grid even if some of those values
happen to look like days or cycles.

Do NOT include a table just because it appears in a section titled "assessments," "activities," or
similar, or because it lists lab test names, procedures, or clinical terms. Specifically exclude:
- reference/definition tables that just list test names or panel contents with no visit columns
  (e.g. "Clinical Laboratory Tests" showing which tests belong to which lab panel)
- questionnaires, checklists, or case-report forms, including ones with Yes/No/Unknown answer
  columns — those columns are answer choices, not study visits, even if the table lives in a
  section about adverse events or assessments
- overview/summary tables that just pair a phase or item name with a paragraph of descriptive text
  (e.g. "Table 1 Overall Trial Plan" pairing "Screening" / "Treatment Period" / "Follow-up" with
  prose explaining each) — there's no repeated visit-column grid here, just narrative text
- per-item dosing or regimen tables where each row is its own independent medication/procedure with
  its own Dose/Start/Stop (e.g. "Mandatory Prophylaxis for Skin Toxicities" listing several drugs,
  each with its own start day and stop day) — these describe individual treatment schedules, not a
  shared visit-by-activity grid
- any other narrative, definitional, or informational table that isn't itself a visit-by-activity
  grid

For each real match, respond on its own line:
SOA_FOUND: <table label> | StartPage: <page> | EndPage: <page, including continuation/footnote pages>
If none: NO_SOA_FOUND"""},
        ]}],
    )
    findings = []
    for line in (response.text or "").splitlines():
        m = re.search(
            r"SOA_FOUND\s*:?\s*(.+?)\s*\|\s*StartPage\s*:?\s*(\d+)\s*\|\s*EndPage\s*:?\s*(\d+)",
            line, re.I,
        )
        if m:
            label, start = m.group(1).strip(), int(m.group(2))
            findings.append((_enrich_label(pdf_path, label, start), start, int(m.group(3))))
    print(f"Found {len(findings)} table(s): {[f[0] for f in findings]}")
    return findings


_GENERIC_LABEL_RE = re.compile(r"^(?:appendix\s+[a-z0-9]+\.?\s*)?table\s+[a-z0-9]+\.?$", re.I)


def _enrich_label(pdf_path, label, page_num):
    """Gemini's label for a table sometimes comes back as just "Table 4."
    with no description, and how verbose it is can vary run to run even for
    the same PDF. Rather than trust that wording, when the label is this
    generic, read the table's actual printed caption straight off the PDF
    page via PyMuPDF — same input, same output every time, and zero extra
    API calls. Leaves any already-descriptive label untouched."""
    if not _GENERIC_LABEL_RE.match(label):
        return label
    ident_match = re.search(r"table\s+([a-z0-9]+)", label, re.I)
    if not ident_match:
        return label
    ident = ident_match.group(1)
    doc = fitz.open(pdf_path)
    if page_num - 1 < 0 or page_num - 1 >= len(doc):
        return label
    text = doc[page_num - 1].get_text()
    m = re.search(rf"table\s+{re.escape(ident)}\.\s+([^\n]+)", text, re.I)
    if not m:
        return label
    desc = m.group(1).strip()
    enriched = f"Table {ident}. {desc}"
    prefix = text[:m.start()]
    am = re.search(r"appendix\s+([a-z0-9]+)\.\s+([^\n]*?)\s*\Z", prefix, re.I)
    if am and len(prefix) - am.start() < 80:
        enriched = f"Appendix {am.group(1)}. {am.group(2).strip()} — {enriched}"
    return enriched


RENDER_DPI = 300


def render_pages_as_images(pdf_path, start_page, end_page):
    doc = fitz.open(pdf_path)
    images = []
    for pg in range(start_page - 1, end_page):
        if pg < len(doc):
            pix = doc[pg].get_pixmap(dpi=RENDER_DPI)
            images.append(pix.tobytes("png"))
    return images


def _clean_header_cell(text):
    return re.sub(r"\s+", " ", (text or "").replace("\n", " ")).strip()


def _clean_data_cell(text):
    """PyMuPDF's table extractor sometimes emits a footnote letter on its own
    line next to the mark (e.g. "c\\nX") because of how superscript text sits
    above the baseline in the PDF's real layout — normalize that into "X^c"
    instead of leaving it as garbled multi-line text."""
    text = (text or "").strip()
    if not text:
        return ""
    m = re.match(r"^([a-zA-Z,]+)\s*\n?\s*[Xx]$", text) or re.match(r"^[Xx]\s*\n?\s*([a-zA-Z,]+)$", text)
    if m:
        return f"X^{m.group(1)}"
    return re.sub(r"\s+", " ", text.replace("\n", " ")).strip()


def extract_table_via_pdf_native(pdf_path, start_page, end_page):
    """Tier 1, tried before any image/vision approach at all: many protocol
    PDFs are born-digital (Word/authoring-tool exports), not scans — their
    tables are real vector-drawn grids with real selectable text, which
    PyMuPDF can parse directly via its built-in table parser. When this
    applies, it's exact by construction (no visual interpretation happens at
    all) and costs zero API calls. Returns (columns, rows) or (None, None) if
    no usable table is found on these pages (e.g. because the table is
    actually a flattened image/scan), so the caller can fall through to the
    image-based tiers instead."""
    doc = fitz.open(pdf_path)
    columns = None
    rows = []
    header_positions = set()
    data_positions = set()
    for pg in range(start_page - 1, min(end_page, len(doc))):
        page = doc[pg]
        try:
            tabs = page.find_tables()
        except Exception:
            continue
        if not tabs.tables:
            continue
        t = max(tabs.tables, key=lambda tt: tt.row_count * tt.col_count)
        if t.row_count < 2 or t.col_count < 2:
            continue
        data = t.extract()
        if not data:
            continue
        header = [_clean_header_cell(c) for c in data[0][1:]]
        body = data[1:]
        if columns is None:
            columns = header
            header_positions.update(i for i, c in enumerate(header) if c)
        elif [h.lower() for h in header] != [c.lower() for c in columns]:
            # This page's table didn't repeat the same header row, so its
            # own first row is actually data, not a header to discard.
            body = data
        for r in body:
            if len(r) - 1 != len(columns):
                continue
            activity = _clean_header_cell(r[0])
            if activity:
                cleaned = [_clean_data_cell(c) for c in r[1:]]
                data_positions.update(i for i, c in enumerate(cleaned) if c)
                rows.append([activity] + cleaned)
    if not columns or not rows:
        return None, None
    if data_positions:
        overlap = len(data_positions & header_positions) / len(data_positions)
        if overlap < 0.5:
            # Some tables with deeply nested/merged multi-row headers (e.g. a
            # "Phase" row spanning several "Day" sub-columns) get parsed by
            # PyMuPDF with the header text and the data marks landing on
            # different raw grid columns -- real, seen in practice, not
            # hypothetical. When most marks don't line up with any column
            # that actually has header text, that's a sign the header and
            # body were read off two different column grids, so the mapping
            # from mark to visit can't be trusted. Bail out here rather than
            # report a mark under the wrong visit, and let the caller fall
            # back to reading the rendered page image instead.
            return None, None
    return columns, rows


def _detect_column_bounds(img):
    """Many Schedule of Activities tables (e.g. AbbVie's template) mark a cell by
    filling it with a solid color rather than drawing a small checkmark glyph on
    a busy grid. Reading that reliably is a color/geometry problem, not something
    that needs a vision model to judge — so instead of asking Gemini to "look
    carefully," this finds the table's actual column grid lines directly from the
    image: the grid lines partition the page into enclosed regions (one per
    cell), found here via connected components on the inverse of the line mask."""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, line_mask = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    line_mask = cv2.dilate(line_mask, kernel, iterations=1)
    cell_mask = cv2.bitwise_not(line_mask)
    num_labels, _, stats, _ = cv2.connectedComponentsWithStats(cell_mask, connectivity=4)
    h, w = gray.shape
    page_area = h * w
    xs = set()
    for i in range(1, num_labels):
        x, y, cw, ch, area = stats[i]
        if area < page_area * 0.0005 or area > page_area * 0.3:
            continue
        if ch > h * 0.2:
            continue
        xs.add(x)
        xs.add(x + cw)
    xs = sorted(xs)
    merged = []
    for x in xs:
        if not merged or x - merged[-1] > 15:
            merged.append(x)
        else:
            merged[-1] = (merged[-1] + x) // 2
    return merged


def _detect_row_bounds(img, x_start, x_end):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    _, bw = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
    h, w = bw.shape
    sub = bw[:, x_start:x_end]
    horiz_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(w // 3, 1), 1))
    horiz = cv2.erode(sub, horiz_kernel, iterations=1)
    horiz = cv2.dilate(horiz, horiz_kernel, iterations=1)
    row_profile = horiz.sum(axis=1) / 255
    min_len = (x_end - x_start) * 0.6
    lines, in_line, start = [], False, 0
    for i, v in enumerate(row_profile):
        if v > min_len and not in_line:
            in_line, start = True, i
        elif v <= min_len and in_line:
            in_line = False
            lines.append((start + i) // 2)
    if in_line:
        lines.append((start + len(row_profile)) // 2)
    return lines


def _classify_cell(img, x0, x1, y0, y1):
    """A cell is 'marked' if a large majority of its full area is filled with
    color other than white — this doesn't hardcode any particular fill color,
    so it isn't tied to AbbVie's specific blue.

    This checks the FULL cell area (minus a thin border margin), not just a
    small sample near the center. A real confirmed bug: on tables with a
    section-divider row (e.g. "INTERVIEWS AND QUESTIONNAIRES"), that row's
    own bold title text is sometimes long enough to visually spill rightward
    underneath the grid lines into the first few checkbox columns. That
    stray text tints maybe 30% of a cell — nowhere near a genuine solid-fill
    mark, which covers ~100% of its cell — but a small centered sample patch
    could land right on top of a stray letter and misread it as marked.
    Requiring most of the cell to be non-white rejects that bleed-through
    while still reliably catching real marks."""
    region = img[y0 + 3:y1 - 3, x0 + 3:x1 - 3]
    if region.size == 0:
        return ""
    flat = region.reshape(-1, 3)
    is_white = (flat[:, 0] > 235) & (flat[:, 1] > 235) & (flat[:, 2] > 235)
    frac_nonwhite = 1 - (is_white.sum() / len(flat))
    return "✓" if frac_nonwhite > 0.6 else ""


def _row_label_text(page, x0_px, x1_px, y0_px, y1_px, dpi):
    """Reads the real Activity-column text straight from the PDF page's own
    text layer, inside this row band's true pixel rectangle (converted back
    to PDF point-space). When available, this is ground truth for row
    IDENTITY, independent of whatever order a separate Gemini call happens to
    list activities in. Many of these tables are flattened/scanned images
    with no text layer at all (e.g. AbbVie's blue-fill style), in which case
    this returns "" and the caller falls back to positional matching."""
    scale = 72.0 / dpi
    rect = fitz.Rect(x0_px * scale, y0_px * scale, x1_px * scale, y1_px * scale)
    return re.sub(r"\s+", " ", page.get_text("text", clip=rect)).strip()


def detect_marks_via_image(pdf_path, start_page, end_page, columns, activities):
    """Deterministic, zero-API-call alternative to asking Gemini to read the
    grid: detect the table's real column/row lines and classify each cell by
    color. Returns None (triggering a fallback to the Gemini-based reader) if
    the detected grid doesn't look sane for this table — e.g. a differently
    styled table that doesn't use solid-fill marks, or a real digital-text
    table this technique was never meant for. Returns a list of
    (raw_activity_text_or_"", marks) pairs, in true visual order."""
    doc = fitz.open(pdf_path)
    all_page_rows = []
    for pg in range(start_page - 1, end_page):
        if pg >= len(doc):
            continue
        page = doc[pg]
        page_bytes = page.get_pixmap(dpi=RENDER_DPI).tobytes("png")
        img = cv2.imdecode(np.frombuffer(page_bytes, dtype=np.uint8), cv2.IMREAD_COLOR)
        if img is None:
            return None
        col_bounds = _detect_column_bounds(img)
        if len(col_bounds) - 2 != len(columns):
            return None
        row_bounds = _detect_row_bounds(img, col_bounds[0], col_bounds[-1])
        page_rows = []
        for i in range(len(row_bounds) - 1):
            y0, y1 = row_bounds[i], row_bounds[i + 1]
            row = [
                _classify_cell(img, col_bounds[c + 1], col_bounds[c + 2], y0, y1)
                for c in range(len(col_bounds) - 2)
            ]
            if any(row):
                label = _row_label_text(page, col_bounds[0], col_bounds[1], y0, y1, RENDER_DPI)
                page_rows.append((label, row))

        # Multi-level column headers (a "Phase" bar spanning several "Day"
        # sub-columns, repeated at the top of every page) are solid-filled
        # the same way a marked cell is, so they show up here as extra
        # fully-"marked" row-bands ahead of the real first activity -- not
        # real data. Strip that leading run before it gets treated as such.
        # A genuine activity row essentially always has at least one blank
        # column, so this only ever removes header artifacts in practice;
        # in the rare case it took a real row too, the length check below
        # still catches it and falls back to Gemini instead of guessing.
        start_idx = 0
        while start_idx < len(page_rows) and all(page_rows[start_idx][1]):
            start_idx += 1
        all_page_rows.extend(page_rows[start_idx:])

    if len(all_page_rows) != len(activities):
        return None
    return all_page_rows


def _call_gemini_json(image_parts, prompt_text, schema):
    response = client.models.generate_content(
        model="gemini-2.5-pro",
        contents=[{"parts": [*image_parts, {"text": prompt_text}]}],
        config=types.GenerateContentConfig(
            response_mime_type="application/json",
            response_json_schema=schema,
            thinking_config=types.ThinkingConfig(thinking_budget=-1),
        ),
    )
    text = (response.text or "").strip()
    return json.loads(text) if text else None


def extract_structure(label, image_parts):
    data = _call_gemini_json(
        image_parts,
        f'This is a table titled "{label}" from a clinical protocol.\n\n{STRUCTURE_INSTRUCTIONS}',
        STRUCTURE_SCHEMA,
    )
    if not data:
        return [], [], []
    columns = [str(c).strip() for c in data.get("columns", [])]
    activities = [str(a).strip() for a in data.get("activities", [])]
    footnotes = [
        (str(f.get("marker", "")).strip(), str(f.get("text", "")).strip())
        for f in data.get("footnotes", [])
        if str(f.get("text", "")).strip()
    ]
    return columns, activities, footnotes


def extract_marks_for_columns(label, image_parts, columns_subset, all_columns, activities):
    if not columns_subset or not activities:
        return {}
    data = _call_gemini_json(
        image_parts,
        f'This is a table titled "{label}" from a clinical protocol.\n\n'
        f'{build_marks_instructions(columns_subset, all_columns, activities)}',
        MARKS_SCHEMA,
    )
    if not data:
        return {}
    marks_by_activity = {}
    for row in data.get("rows", []):
        act = str(row.get("activity", "")).strip()
        marks_by_activity[act] = row.get("marked_cells", [])
    return marks_by_activity


def _normalize(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _lookup_marks(activity, norm_to_marks_key, marks_by_activity):
    norm = _normalize(activity)
    key = norm_to_marks_key.get(norm)
    if key is None:
        close = difflib.get_close_matches(norm, list(norm_to_marks_key.keys()), n=1, cutoff=0.8)
        key = norm_to_marks_key.get(close[0]) if close else None
    return marks_by_activity.get(key, []) if key else []


def extract_table(label, pdf_path, start_page, end_page):
    native_columns, native_rows = extract_table_via_pdf_native(pdf_path, start_page, end_page)
    if native_columns and native_rows:
        print(f"  '{label}': read directly from the PDF's real table data (0 API calls, exact by construction).")
        return [["Activity"] + native_columns] + native_rows, []

    image_parts = [
        {"inline_data": {"mime_type": "image/png", "data": img}}
        for img in render_pages_as_images(pdf_path, start_page, end_page)
    ]

    columns, activities, footnotes = extract_structure(label, image_parts)
    if not columns or not activities:
        return [], []

    image_rows = detect_marks_via_image(pdf_path, start_page, end_page, columns, activities)
    if image_rows is not None:
        print(f"  '{label}': marks read via deterministic image detection (0 extra API calls).")
        norm_to_activity = {_normalize(a): a for a in activities}
        rows = [["Activity"] + columns]
        for i, (raw_label, marks) in enumerate(image_rows):
            matched = None
            if raw_label:
                # This table has a real text layer (not a flattened/scanned
                # page) -- use the row's own on-page text to identify which
                # activity it is, rather than trusting position alone.
                norm = _normalize(raw_label)
                matched = norm_to_activity.get(norm)
                if matched is None:
                    close = difflib.get_close_matches(norm, list(norm_to_activity.keys()), n=1, cutoff=0.6)
                    matched = norm_to_activity.get(close[0]) if close else None
            # No text layer to check against (e.g. a scanned page): fall
            # back to positional order, same as before this change.
            fallback = activities[i] if i < len(activities) else raw_label
            rows.append([matched or raw_label or fallback] + marks)
        return rows, footnotes

    print(f"  '{label}': image-based grid detection didn't produce a clean match "
          f"(different table style than expected) — falling back to Gemini for marks.")
    marks_by_activity = extract_marks_for_columns(label, image_parts, columns, columns, activities)
    norm_to_marks_key = {_normalize(k): k for k in marks_by_activity}

    rows = [["Activity"] + columns]
    for activity in activities:
        marks = [""] * len(columns)
        for cell in _lookup_marks(activity, norm_to_marks_key, marks_by_activity):
            col = str(cell.get("column", "")).strip()
            if col in columns:
                marks[columns.index(col)] = str(cell.get("value", "")).strip() or "✓"
        rows.append([activity] + marks)

    return rows, footnotes


def save_to_excel(findings, pdf_path, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used = set()
    for label, start, end in findings:
        name = re.sub(r'[\\/*?:\[\]]', "", label)[:28] or f"SoA_{start}-{end}"
        n, i = name, 2
        while n in used:
            n, i = f"{name}_{i}", i + 1
        used.add(n)
        ws = wb.create_sheet(title=n)

        rows, footnotes = extract_table(label, pdf_path, start, end)
        ws.append([f"{label} (Pages {start}-{end})"])
        if rows:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(rows[0]))
            for r in rows:
                ws.append(r)
            for cell in ws[2]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
            ws.column_dimensions["A"].width = 45

            if footnotes:
                ws.append([])
                fn_header_row = ws.max_row + 1
                ws.append(["Footnotes"])
                ws.cell(row=fn_header_row, column=1).font = Font(bold=True)
                for marker, text in footnotes:
                    ws.append([f"{marker}: {text}"])

            print(f"  {label}: {len(rows) - 1} row(s), {len(footnotes)} footnote(s) extracted")
        else:
            ws.append(["No data extracted from these pages."])
            print(f"  {label}: FAILED — no data extracted")
    wb.save(output_path)
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    pdf_path = input("Enter path to protocol PDF: ").strip()
    findings = detect_soas(pdf_path)
    if findings:
        save_to_excel(findings, pdf_path, "soa_tables.xlsx")
    else:
        print("No SoAs found.")